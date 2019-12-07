from openpyxl import load_workbook
from os.path import join, exists
from os import remove
from constant import ALLOWED_EXTENSIONS


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def is_xlsx(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == "xlsx"


HEADERS = [
    'ANO_APLICACION',
    'PERIODO_APLICACION',
    'NOMBRE_UA_CURSO',
    'NOMBRE_ASIGNATURA',
    'SIGLA',
    'SECCION',
    'NOMBRE_DEL_DOCENTE'
]

def parse_question(sheet, question, headers_position, data):
    headers = HEADERS.copy()
    headers.append("PREGUNTA_{}".format(question))
    max_row = sheet.max_row
    for i in range(2, max_row + 1):
        line = []
        if sheet.cell(row=i, column=1).value is None:
            continue
        for column in headers:
            line.append(sheet.cell(row=i, column=headers_position[column]).value)

        line.append(18)
        data.append([str(x).replace("\n", " ").replace("\r", "").replace("\t", " ") for x in line])
    return data

def check(headers):
    for elem in HEADERS:
        if elem not in headers:
            return False
    return True

def process_encuestas(filename):
    filepath = join("data", filename)
    wb = load_workbook(filepath)
    data = []
    try: 
        sheet = wb['Pregunta 18']
    except:
        sheet = wb['PREGUNTA 18']

    headers = {}
    for j in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=j).value.replace("\n", " ").replace("\r", "").replace("\t", " ")
        headers[header] = j
    print(headers)
    if not check(headers):
        remove(filepath)
        return "ERROR", False

    data = parse_question(sheet, 18, headers, data)
    
    try: 
        sheet = wb['Pregunta 17']
    except:
        sheet = wb['PREGUNTA 17']


    headers = {}
    for j in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=j).value.replace("\n", " ").replace("\r", "").replace("\t", " ")
        headers[header] = j

    if not check(headers):
        remove(filepath)
        return "ERROR", False

    data = parse_question(sheet, 17, headers, data)

    final_header = HEADERS.copy()
    final_header.append("TEXTO")
    final_header.append("PREGUNTA")

    exist = exists(join("data", "EncuestasDocentes.tsv"))
    if exist:
        with open(join("data", "EncuestasDocentes.tsv"), "a", encoding="UTF-8") as file:
            for line in data:
                file.write("\t".join(line) + "\n")

    else:
        with open(join("data", "EncuestasDocentes.tsv"), "w", encoding="UTF-8") as file:
            file.write("\t".join(final_header) + "\n")
            for line in data:
                file.write("\t".join(line) + "\n")

    remove(filepath)
    return "EncuestasDocentes.tsv", exist
