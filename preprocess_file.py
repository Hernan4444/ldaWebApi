from openpyxl import load_workbook
from os.path import join
from os import remove
import pandas as pd
from nltk.tokenize import word_tokenize
from nltk.stem import SnowballStemmer
from string import punctuation
from collections import defaultdict

NON_LETTERS = list(punctuation)

# we add spanish punctuation
NON_LETTERS.extend(['¿', '¡', '-'])
NON_LETTERS.extend(map(str, range(10)))

HEADERS = [
    'ANO_APLICACION',
    'PERIODO_APLICACION',
    'NOMBRE_UA_CURSO',
    'NOMBRE_ASIGNATURA',
    'SIGLA',
    'SECCION',
    'NOMBRE_DEL_DOCENTE'
]

# based on http://www.cs.duke.edu/courses/spring14/compsci290/assignments/lab02.html
STEMMER = SnowballStemmer('spanish')


def remove_stopword(text, stopwords):
    # remove non letters
    text = ''.join([c.lower() for c in text if c.lower() not in NON_LETTERS])

    # remove stopword
    text = ' '.join([c.lower() for c in text.split(" ") if c.lower() not in stopwords])

    # tokenize
    tokens = word_tokenize(text)
    if len(tokens):
        return " ".join(tokens)
    return " "


def stemming_text(text):
    new_words = [STEMMER.stem(word) for word in word_tokenize(text)]
    if len(new_words):
        return " ".join(new_words)
    return " "


def generate_new_text(df):
    stopwords = [x.strip() for x in open("stopwords_spanish.txt", encoding="UTF-8")]

    df['TEXTO_STOPWORD'] = df.TEXTO.map(lambda text: remove_stopword(text, stopwords))
    df['TEXTO_STEMMING'] = df.TEXTO.map(lambda text: stemming_text(text))
    df['TEXTO_STOPWORD_STEMMING'] = df.TEXTO_STOPWORD.map(lambda text: stemming_text(text))
    return df


def parse_question(sheet, question, headers_position, data):
    headers = HEADERS.copy()
    headers.append("PREGUNTA_{}".format(question))
    max_row = sheet.max_row
    for i in range(2, max_row + 1):
        line = []
        if sheet.cell(row=i, column=1).value is None:
            continue
        for column in headers:
            value = sheet.cell(row=i, column=headers_position[column]).value
            if str(value) == "None":
                value = "Sin texto"
            line.append(value)

        line.append(question)
        data.append([str(x).replace("\n", " ").replace("\r", "").replace("\t", " ") for x in line])
    return data


def check(headers):
    for elem in HEADERS:
        if elem not in headers:
            return False
    return True


def index_file(df):
    indexs = defaultdict(list)
    for index, text in enumerate(df.TEXTO_STEMMING.values):
        # remove non letters
        text = ''.join([c.lower() for c in text if c.lower() not in NON_LETTERS])

        # tokenize
        tokens = word_tokenize(text)
        for word in tokens:
            indexs[word].append(index)
    return indexs


def process_encuestas(filename, exist_filename):
    filepath = join("data", filename)
    workbook = load_workbook(filepath)

    data = []
    try:
        sheet = workbook['Pregunta 18']
    except:
        sheet = workbook['PREGUNTA 18']

    headers = {}
    for j in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=j).value.replace("\n", " ").replace("\r", "").replace("\t", " ")
        headers[header] = j

    print(headers)
    if not check(headers):
        remove(filepath)
        return None, ""

    data = parse_question(sheet, 18, headers, data)

    try:
        sheet = workbook['Pregunta 17']
    except:
        sheet = workbook['PREGUNTA 17']

    headers = {}
    for j in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=j).value.replace(
            "\n", " ").replace("\r", "").replace("\t", " ")
        headers[header] = j

    if not check(headers):
        remove(filepath)
        return None, ""

    data = parse_question(sheet, 17, headers, data)

    final_header = HEADERS.copy()
    final_header.append("TEXTO")
    final_header.append("PREGUNTA")
    df = pd.DataFrame(data, columns=final_header)
    df['SEMESTRE'] = df.PERIODO_APLICACION.map(lambda x: 1 if x == 20 else 2)
    del df['PERIODO_APLICACION']
    df = generate_new_text(df)
    indexs = index_file(df)

    remove(filepath)
    if exist_filename != "":
        last_dataset = pd.read_csv(join('data', exist_filename), sep="\t", index_col=0)
        df = pd.concat([df, last_dataset])

    return df, indexs


def process_other_text(filename):
    filepath = join("data", filename)
    with open(filepath, encoding="UTF-8") as file:
        texts = file.readlines()
    df = pd.DataFrame([texts], columns="TEXTO")
    df = generate_new_text(df)
    indexs = index_file(df)
    remove(filepath)
    return df, indexs
